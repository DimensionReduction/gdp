import requests
import parsel
import openpyxl
import pandas as pd

headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/98.0.4758.80 Safari/537.36',
           'sec-ch-ua': '" Not A;Brand";v="99", "Chromium";v="98", "Google Chrome";v="98"'}

wb = openpyxl.Workbook('GDP.xlsx')
for i in range(1980, 2022):
    url = 'http://www.8pu.com/gdp/ranking_' + str(i) + '.html'
    response = requests.get(url, headers=headers)
    response.encoding = response.apparent_encoding
    selector = parsel.Selector(response.text)
    lis = selector.css('#table01')
    for li in lis:
        rank = li.css('#US_ > td.rank > font > font::text').getall()
        rank = ','.join(rank).replace('排名第', '').split(',')
        country = li.css('#US_ > td:nth-child(2) > a > font::text').getall()
        dollar = li.css('#US_ > td.value > font::text').getall()
        dollar = ','.join(dollar).replace('$', '').split(',')
        rmb = li.css('#US_ > td.rank_prev > font::text').getall()
        rmb = ','.join(rmb).replace('￥', '').replace('亿元', '').split(',')
        continent = li.css('#US_ > td.area > font > font::text').getall()
        continent = ','.join(continent).replace('国家', '').replace('州','洲').split(',')
        dict = {'排名': rank, '国家/地区': country, 'GDP总量(美元核算)': dollar, 'GDP总量(人民币核算)': rmb, '所属洲': continent}
    d = list(zip(rank, country, dollar, rmb, continent))
    print('\n//////////  ' + str(i) + '年全球GDP  //////////\n')
    # for x1, x2, x3, x4, x5 in d:
    #     print(x1, x2, x3, x4, x5)
    name=[[u'排名',u'国家/地区',u'GDP总量(美元核算)',u'GDP总量(人民币核算)',u'所属洲']]
    sheet=wb.create_sheet(index=0,title=str(i))
    for r in name:
        sheet.append(r)
    for rr in d:
        sheet.append(rr)
wb.save('GDP.xlsx')

wbb =openpyxl.load_workbook('GDP.xlsx')
sheet_names=wbb.sheetnames
for ii in sheet_names:
    ws=wbb[ii]
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
wbb.save('GDP.xlsx')
